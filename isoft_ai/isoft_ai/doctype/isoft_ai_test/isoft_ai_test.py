import frappe
import openai
import re
import json
import csv
import io
import os
from typing import List, Dict, Optional
import pdfkit
try:
    import pdfkit
except ImportError:
    pdfkit = None
from frappe.model.document import Document
from frappe.utils import escape_html
try:
    import sqlparse
except ImportError:
    sqlparse = None
import difflib
import hashlib
from datetime import datetime, timedelta

# Cache settings
CACHE_EXPIRY_HOURS = 24
MAX_CACHE_ENTRIES = 1000

# Dynamic cache expiry based on query type
CACHE_EXPIRY_RULES = {
    'REAL_TIME': 0,      # No caching for real-time data
    'HIGH_FREQ': 5,      # 5 minutes for frequently changing data
    'MEDIUM_FREQ': 30,   # 30 minutes for moderately changing data
    'LOW_FREQ': 120,     # 2 hours for slowly changing data
    'STATIC': 1440       # 24 hours for static/reference data
}

# ERPNext v13 Module Coverage
ERPNEXT_MODULES = {
    'ACCOUNTING': ['GL Entry', 'Journal Entry', 'Payment Entry', 'Account'],
    'SELLING': ['Sales Invoice', 'Quotation', 'Customer'],
    'BUYING': ['Purchase Invoice', 'Purchase Receipt', 'Supplier'],
    'STOCK': ['Stock Entry', 'Stock Ledger Entry', 'Bin', 'Item', 'Warehouse'],
    'MANUFACTURING': ['Work Order', 'BOM', 'Production Plan', 'Job Card'],
    'HR': ['Employee', 'Salary Slip', 'Attendance', 'Leave Application'],
    'PROJECTS': ['Project', 'Task', 'Timesheet', 'Issue'],
    'CRM': ['Lead', 'Opportunity', 'Customer', 'Contact'],
    'ASSETS': ['Asset', 'Asset Movement', 'Asset Maintenance'],
    'QUALITY': ['Quality Inspection', 'Quality Goal']
}

def get_cache_key(question: str, chat_history: Optional[List] = None) -> str:
    """Generate cache key for question and context"""
    content = question
    if chat_history:
        # Include last 2 messages for context
        recent_context = str(chat_history[-2:]) if len(chat_history) > 2 else str(chat_history)
        content += recent_context
    return hashlib.md5(content.encode()).hexdigest()

def determine_cache_expiry(question: str, intent: str, suggested_doctypes: list) -> int:
    """Determine appropriate cache expiry based on query characteristics"""
    
    question_lower = question.lower()
    
    # Real-time indicators - NO CACHING
    real_time_keywords = [
        'current', 'now', 'today', 'latest', 'recent', 'live', 'real-time',
        'current stock', 'current balance', 'today\'s', 'this month',
        'pending', 'open', 'active', 'running', 'ongoing'
    ]
    
    if any(keyword in question_lower for keyword in real_time_keywords):
        return CACHE_EXPIRY_RULES['REAL_TIME']
    
    # High frequency changing data - 5 minutes
    high_freq_keywords = [
        'stock', 'inventory', 'balance', 'quantity', 'amount', 'total',
        'sales', 'purchase', 'invoice', 'payment', 'transaction',
        'delivery', 'receipt', 'movement', 'entry'
    ]
    
    if any(keyword in question_lower for keyword in high_freq_keywords):
        return CACHE_EXPIRY_RULES['HIGH_FREQ']
    
    # Medium frequency - 30 minutes
    medium_freq_keywords = [
        'customer', 'supplier', 'employee', 'project', 'task',
        'report', 'summary', 'analysis', 'performance', 'status'
    ]
    
    if any(keyword in question_lower for keyword in medium_freq_keywords):
        return CACHE_EXPIRY_RULES['MEDIUM_FREQ']
    
    # Low frequency - 2 hours
    low_freq_keywords = [
        'master', 'setup', 'configuration', 'settings', 'reference',
        'item group', 'warehouse', 'account', 'cost center'
    ]
    
    if any(keyword in question_lower for keyword in low_freq_keywords):
        return CACHE_EXPIRY_RULES['LOW_FREQ']
    
    # Static/reference data - 24 hours
    static_keywords = [
        'what is', 'how to', 'explain', 'definition', 'meaning',
        'knowledge', 'help', 'guide', 'tutorial'
    ]
    
    if any(keyword in question_lower for keyword in static_keywords):
        return CACHE_EXPIRY_RULES['STATIC']
    
    # Default based on intent
    if intent in ['KNOWLEDGE', 'CLARIFY']:
        return CACHE_EXPIRY_RULES['STATIC']
    elif intent in ERPNEXT_MODULES:
        return CACHE_EXPIRY_RULES['MEDIUM_FREQ']
    
    # Conservative default
    return CACHE_EXPIRY_RULES['HIGH_FREQ']

def get_cached_response(cache_key: str) -> Optional[Dict]:
    """Get cached response if available and not expired"""
    try:
        # Check if document exists first
        if frappe.db.exists('AI Cache', cache_key):
            cache_doc = frappe.get_doc('AI Cache', cache_key)
            if cache_doc.expires_at > datetime.now():
                return json.loads(cache_doc.response_data)
    except Exception as e:
        frappe.logger().debug(f"Cache get error (normal): {str(e)}")
    return None

def set_cached_response(cache_key: str, response_data: dict, expiry_hours: int = CACHE_EXPIRY_HOURS):
    """Cache response with expiry"""
    try:
        expires_at = datetime.now() + timedelta(hours=expiry_hours)
        
        # Check if cache already exists
        if frappe.db.exists('AI Cache', cache_key):
            try:
                cache_doc = frappe.get_doc('AI Cache', cache_key)
                cache_doc.response_data = json.dumps(response_data)
                cache_doc.expires_at = expires_at
                cache_doc.save(ignore_permissions=True)
            except Exception as e:
                frappe.logger().debug(f"Cache update error: {str(e)}")
        else:
            # Create new cache entry
            try:
                cache_doc = frappe.get_doc({
                    'doctype': 'AI Cache',
                    'name': cache_key,
                    'response_data': json.dumps(response_data),
                    'expires_at': expires_at
                })
                cache_doc.insert(ignore_permissions=True)
            except Exception as e:
                frappe.logger().debug(f"Cache create error: {str(e)}")
        
        # Clean old cache entries if too many
        cleanup_old_cache()
    except Exception as e:
        frappe.logger().debug(f"Cache set error: {str(e)}")

def cleanup_old_cache():
    """Remove expired cache entries and limit total entries"""
    try:
        # Check if table exists first
        if not frappe.db.exists('DocType', 'AI Cache'):
            return
            
        # Remove expired entries
        frappe.db.sql("DELETE FROM `tabAI Cache` WHERE expires_at < %s", (datetime.now(),))
        
        # Limit total entries
        total_count = frappe.db.count('AI Cache')
        if total_count > MAX_CACHE_ENTRIES:
            excess = total_count - MAX_CACHE_ENTRIES
            frappe.db.sql("""
                DELETE FROM `tabAI Cache` 
                ORDER BY creation ASC 
                LIMIT %s
            """, (excess,))
    except Exception as e:
        frappe.logger().debug(f"Cache cleanup error: {str(e)}")

def generate_clarifying_question(question: str, chat_history: list, suggested_doctypes: list, token_usage: dict) -> str:
    """Generate a helpful clarifying question"""
    context = ""
    if suggested_doctypes:
        context = f"I detected this might be related to: {', '.join(suggested_doctypes)}. "
    
    recent_context = ""
    if chat_history and len(chat_history) > 0:
        recent_context = f"Looking at our conversation, "
    
    prompt = [
        {
            "role": "system", 
            "content": (
                f"You are a helpful ERPNext v13 assistant. The user's question needs clarification. "
                f"Generate a friendly, specific clarifying question that helps narrow down what they want. "
                f"Use emojis sparingly and be conversational. Available ERPNext modules: {list(ERPNEXT_MODULES.keys())}"
            )
        },
        {
            "role": "user", 
            "content": f"User question: '{question}'\nContext: {context}{recent_context}\nSuggested modules: {suggested_doctypes}"
        }
    ]
    
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=prompt,
        max_tokens=100,
        temperature=0.3,
    )
    
    token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += response['usage']['completion_tokens']
    token_usage["total_tokens"] += response['usage']['total_tokens']
    
    return response.choices[0].message["content"].strip()

def handle_erpnext_module_query(intent: str, question: str, suggested_doctypes: list, confidence: float, token_usage: dict) -> str:
    """Handle specific ERPNext module queries with enhanced functionality"""
    
    if intent in ERPNEXT_MODULES and confidence > 0.7:
        # Generate module-specific response
        module_doctypes = ERPNEXT_MODULES[intent]
        relevant_types = [dt for dt in suggested_doctypes if dt in module_doctypes] or module_doctypes[:3]
        
        try:
            # Try to generate SQL for the query
            sql_query = generate_enhanced_sql(question, intent, relevant_types, token_usage)
            if sql_query:
                db_result = frappe.db.sql(sql_query, as_dict=True)
                if db_result:
                    if len(db_result) > 10 or len(db_result[0].keys()) > 5:
                        return generate_excel_file(db_result)
                    else:
                        return polish_erp_answer_html(question, format_result(db_result), token_usage)
                else:
                    return f"<div class='alert alert-info'>🔍 No data found for your {intent.lower()} query. Try adjusting your criteria or time range.</div>"
            else:
                return f"<div class='alert alert-warning'>⚠️ Could not generate a query for this {intent.lower()} request. Please be more specific about what data you need.</div>"
                
        except Exception as e:
            return f"<div class='alert alert-danger'>❌ Error processing {intent.lower()} query: {str(e)}</div>"
    
    return f"<div class='alert alert-info'>🚧 {intent} module functionality is being enhanced. Please try a more specific query.</div>"

def generate_enhanced_sql(question: str, intent: str, suggested_doctypes: list, token_usage: dict) -> Optional[str]:
    """Enhanced SQL generation with ERPNext v13 module-specific knowledge"""
    
    module_context = ""
    if intent in ERPNEXT_MODULES:
        relevant_doctypes = ERPNEXT_MODULES[intent]
        module_context = f"Focus on {intent} module with doctypes: {relevant_doctypes}. "
    
    if suggested_doctypes:
        module_context += f"Prioritize these doctypes: {suggested_doctypes}. "
    
    messages = [
        {
            "role": "system",
            "content": (
                f"You are an expert ERPNext v13 SQL generator. {module_context}"
                f"Generate ONLY a valid SELECT query for ERPNext v13 with 'tab' prefixed tables.\n\n"
                f"ERPNext v13 Schema Rules:\n"
                f"- All tables prefixed with 'tab' and use spaces (e.g., 'tabSales Invoice', 'tabSales Invoice Item')\n"
                f"- Use docstatus=1 for submitted documents\n"
                f"- Use is_cancelled=0 where applicable\n"
                f"- Parent-child relations use 'parent' field\n"
                f"- Common fields: name, creation, modified, owner\n\n"
                f"ACCOUNTING MODULE:\n"
                f"- tabSales Invoice: customer, posting_date, grand_total, docstatus, outstanding_amount, due_date\n"
                f"- tabPayment Entry: party, posting_date, paid_amount, reference_no, reference_date\n"
                f"- tabGL Entry: posting_date, account, debit, credit, voucher_type, party\n\n"
                f"STOCK MODULE:\n"
                f"- tabItem: item_code, item_name, item_group, brand, description\n"
                f"- tabSales Invoice Item: item_code, qty, rate, amount, parent\n"
                f"- tabSales Invoice: customer, posting_date, grand_total, docstatus\n\n"
                f"SELLING MODULE:\n"
                f"- tabSales Invoice: customer, posting_date, grand_total, docstatus, outstanding_amount\n"
                f"- tabCustomer: customer_name, customer_type, territory\n\n"
                f"Available modules: {ERPNEXT_MODULES}\n\n"
                f"FOLLOW-UP CONTEXT: If the question mentions 'include outstanding', 'add totals', 'show amounts', etc., make sure to include outstanding_amount, grand_total, or relevant financial fields.\n\n"
                f"CRITICAL: You must return ONLY a valid SQL SELECT query. Do NOT provide explanations, instructions, or any other text.\n"
                f"For 'top N' queries, use ORDER BY and LIMIT clauses.\n"
                f"For date filtering, use YEAR(), MONTH() functions or date comparisons.\n"
                f"For overdue invoices with outstanding: SELECT customer, name, posting_date, grand_total, outstanding_amount, due_date FROM `tabSales Invoice` WHERE docstatus = 1 AND outstanding_amount > 0 AND due_date < CURDATE()\n\n"
                f"Return ONLY the SQL query, no explanations."
            )
        },
        {"role": "user", "content": f"Generate SQL for: {question}"}
    ]
    
    try:
        response = openai.ChatCompletion.create(
            model="gpt-4",
            messages=messages,
            max_tokens=300,
            temperature=0,
        )
        
        token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
        token_usage["completion_tokens"] += response['usage']['completion_tokens']
        token_usage["total_tokens"] += response['usage']['total_tokens']
        
        sql = response.choices[0].message['content'].strip().rstrip(";")
        
        if not sql.lower().startswith("select") or re.search(r";|insert|update|delete|drop|alter|truncate", sql, re.I):
            return None
            
        # Enhanced validation
        validation_error = validate_sql_fields(sql)
        if validation_error:
            frappe.logger().error(f"SQL validation error: {validation_error}")
            return None
            
        return sql
        
    except Exception as e:
        frappe.logger().error(f"SQL generation error: {str(e)}")
        return None

def ask_enhanced_knowledge_question(chat_history: list, question: str, intent: str, confidence: float, token_usage: dict) -> str:
    """Enhanced knowledge question handler with ERPNext v13 context"""
    
    # Build context-aware prompt
    context_info = ""
    if intent != "KNOWLEDGE":
        context_info = f"The user seems to be asking about {intent} (confidence: {confidence:.1f}). "
    
    # Limit chat history for efficiency
    recent_history = chat_history[-4:] if len(chat_history) > 4 else chat_history
    
    system_prompt = (
        f"You are ISOFT ERP AI assistant specialized in ERPNext v13. {context_info}"
        f"Provide helpful, accurate responses about ERPNext functionality, business processes, and ERP concepts. "
        f"Use emojis sparingly for friendliness. Format responses in HTML with proper styling. "
        f"If the question is about specific ERPNext features, explain them in business context. "
        f"Available ERPNext v13 modules: {', '.join(ERPNEXT_MODULES.keys())}. "
        f"Keep responses concise but informative. Never mention 'ERPNext v13' directly - just say 'our ERP system'."
    )
    
    messages = [{"role": "system", "content": system_prompt}]
    
    # Add recent chat history
    for msg in recent_history:
        if msg.get("role") in ("user", "assistant"):
            messages.append({"role": msg["role"], "content": msg["content"]})
    
    messages.append({"role": "user", "content": question})
    
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        max_tokens=500,
        temperature=0.3,
    )
    
    token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += response['usage']['completion_tokens']
    token_usage["total_tokens"] += response['usage']['total_tokens']
    
    return response.choices[0].message['content'].strip()


class ISOFTAITEST(Document):
    pass


@frappe.whitelist()
def ask_ai(user_question: str, chat_history_json: str = "[]", ai_chat_name: str = "") -> dict:
    if not user_question or not user_question.strip():
        return {"ai_response": "<div class='alert alert-warning'>💬 Please ask me something! I'm here to help with your ERPNext queries.</div>", "chat_name": None}

    if "AI User" not in frappe.get_roles(frappe.session.user):
        return {"ai_response": "<div class='alert alert-danger'>🚫 Access denied. Please contact your administrator for AI access permissions.</div>", "chat_name": None}

    api_key = frappe.conf.get("openai_api_key")
    if not api_key:
        frappe.throw("OpenAI API key not configured in site_config.json")

    openai.api_key = api_key
    user_question = user_question.strip()

    try:
        chat_history = json.loads(chat_history_json)
    except Exception:
        chat_history = []

    # Check cache first for similar questions
    cache_key = get_cache_key(user_question, chat_history)
    cached_response = get_cached_response(cache_key)
    if cached_response:
        frappe.logger().info(f"Cache hit for question: {user_question[:50]}...")
        return cached_response

    # Find the first user message for title generation
    first_user_message = None
    for msg in chat_history:
        if msg.get("role") == "user" and msg.get("content") and msg["content"].strip():
            first_user_message = msg["content"].strip()
            break
    if not first_user_message:
        first_user_message = user_question

    ai_chat = get_or_create_ai_chat(ai_chat_name or "", first_user_message or "")

    token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}

    user_question = preprocess_question(user_question, chat_history)
    # --- ENHANCED INTENT DETECTION WITH ERPNext v13 MODULES ---
    intent_and_action_prompt = [
        {
            "role": "system",
            "content": (
                f"You are an ERPNext v13 AI assistant. Analyze the user's question and respond with JSON containing:\n"
                f"1. 'intent': One of {list(ERPNEXT_MODULES.keys())} + ['KNOWLEDGE', 'CLARIFY', 'STUDY', 'GENERAL_REPORT']\n"
                f"2. 'confidence': 0.0-1.0\n"
                f"3. 'suggested_doctypes': Relevant ERPNext doctypes (if applicable)\n"
                f"4. 'requires_sql': true/false if data query needed\n"
                f"5. 'clarification_needed': true/false if question is ambiguous\n\n"
                f"ERPNext v13 modules: {ERPNEXT_MODULES}\n\n"
                f"IMPORTANT: Questions asking for data, reports, lists, counts, totals, rankings, or analytics should be classified as module-specific intents with requires_sql=true.\n\n"
                f"STUDY INTENT: Use 'STUDY' for ANY request that asks for analysis, study, investigation, or detailed examination of specific items, customers, products, or business entities. This includes but is not limited to:\n"
                f"- 'make a study on/for/about/of'\n"
                f"- 'analyze this item/customer/product'\n"
                f"- 'study this item/customer/product'\n"
                f"- 'investigate this item/customer/product'\n"
                f"- 'examine this item/customer/product'\n"
                f"- 'research this item/customer/product'\n"
                f"- 'detailed analysis of'\n"
                f"- 'comprehensive study of'\n"
                f"- 'business analysis of'\n"
                f"- 'performance analysis of'\n"
                f"- 'market analysis of'\n"
                f"- 'sales analysis of'\n"
                f"- 'inventory analysis of'\n"
                f"- 'customer analysis of'\n"
                f"- 'product analysis of'\n"
                f"- 'financial analysis of'\n"
                f"- 'trend analysis of'\n"
                f"- 'report on this item/customer/product'\n"
                f"- 'insights about this item/customer/product'\n"
                f"- 'tell me about this item/customer/product'\n"
                f"- 'what can you tell me about this item/customer/product'\n"
                f"- 'give me details about this item/customer/product'\n"
                f"- 'show me information about this item/customer/product'\n"
                f"- 'provide analysis of this item/customer/product'\n"
                f"- 'generate report for this item/customer/product'\n"
                f"- 'create study for this item/customer/product'\n"
                f"- 'prepare analysis for this item/customer/product'\n"
                f"- 'build report for this item/customer/product'\n"
                f"- 'develop study for this item/customer/product'\n\n"
                f"CRITICAL STUDY PATTERNS: Any question that implies analysis, investigation, or detailed examination of a specific entity (item code, customer name, product identifier) should be classified as STUDY intent with high confidence (0.9+). Look for context clues like 'this', 'that', 'item', 'customer', 'product' followed by identifiers.\n\n"
                f"FOLLOW-UP QUESTIONS: If the question appears to be a follow-up (like 'include outstanding', 'add customers', 'show totals'), treat it as a data query that requires SQL generation.\n\n"
                f"Examples:\n"
                f"'Show me top customers' -> {{'intent':'SELLING','confidence':0.9,'suggested_doctypes':['Customer','Sales Invoice'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'Top 10 items this year' -> {{'intent':'STOCK','confidence':0.9,'suggested_doctypes':['Item','Sales Invoice Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'Sales report for March' -> {{'intent':'SELLING','confidence':0.9,'suggested_doctypes':['Sales Invoice'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'make a study on this item 000125' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'analyze this item CBL-SN4' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'investigate product 1E7D7AA' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'tell me about customer ABC Corp' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Customer'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'what can you tell me about item XYZ-123' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'give me details about this product' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'show me information about customer John Doe' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Customer'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'provide analysis of item ABC123' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'generate report for product XYZ' -> {{'intent':'STUDY','confidence':0.9,'suggested_doctypes':['Item'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'include outstanding' -> {{'intent':'ACCOUNTING','confidence':0.8,'suggested_doctypes':['Sales Invoice','Payment Entry'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'add customers' -> {{'intent':'SELLING','confidence':0.8,'suggested_doctypes':['Customer','Sales Invoice'],'requires_sql':true,'clarification_needed':false}}\n"
                f"'What is an invoice?' -> {{'intent':'KNOWLEDGE','confidence':0.8,'suggested_doctypes':[],'requires_sql':false,'clarification_needed':false}}\n"
                f"'Tell me more' -> {{'intent':'CLARIFY','confidence':0.9,'suggested_doctypes':[],'requires_sql':false,'clarification_needed':true}}"
            )
        },
        {"role": "user", "content": f"Question: {user_question}\nChat context: {str(chat_history[-2:]) if len(chat_history) > 1 else 'None'}"}
    ]

    intent_response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=intent_and_action_prompt,
        max_tokens=150,
        temperature=0,
    )

    token_usage["prompt_tokens"] += intent_response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += intent_response['usage']['completion_tokens']
    token_usage["total_tokens"] += intent_response['usage']['total_tokens']

    intent_analysis = {}
    try:
        intent_analysis = json.loads(intent_response.choices[0].message["content"])
        intent = intent_analysis.get("intent", "KNOWLEDGE")
        confidence = intent_analysis.get("confidence", 0.5)
        suggested_doctypes = intent_analysis.get("suggested_doctypes", [])
        requires_sql = intent_analysis.get("requires_sql", False)
        clarification_needed = intent_analysis.get("clarification_needed", False)
    except Exception as e:
        frappe.logger().error(f"Intent detection failed: {str(e)}")
        intent = "KNOWLEDGE"
        confidence = 0.5
        suggested_doctypes = []
        requires_sql = False
        clarification_needed = False

    frappe.logger().info(f"Detected intent: {intent} (confidence: {confidence}) for question: {user_question}")
    frappe.logger().info(f"Intent analysis: {intent_analysis}")
    frappe.logger().info(f"Requires SQL: {requires_sql}, Suggested doctypes: {suggested_doctypes}")
    
    if intent == "CLARIFY" or clarification_needed:
        clarifying_response = generate_clarifying_question(user_question, chat_history, suggested_doctypes, token_usage)
        add_ai_message(ai_chat, user_question, clarifying_response, token_usage)
        result_data = {"ai_response": clarifying_response, "chat_name": ai_chat.name}
        # No caching for clarification questions
        return result_data

    # Handle ERPNext module-specific queries
    elif intent in ERPNEXT_MODULES or intent == "GENERAL_REPORT":
        try:
            result = handle_erpnext_module_query(intent, user_question, suggested_doctypes, confidence, token_usage)
            add_ai_message(ai_chat, user_question, result, token_usage)
            result_data = {"ai_response": result, "chat_name": ai_chat.name}
            
            # Smart caching based on data volatility
            cache_expiry = determine_cache_expiry(user_question, intent, suggested_doctypes)
            if cache_expiry > 0:  # Only cache if expiry > 0
                set_cached_response(cache_key, result_data, cache_expiry)
            
            return result_data
        except Exception as e:
            msg = str(e)
            # Enhanced error handling with suggestions
            if 'does not exist' in msg or 'Unknown column' in msg:
                all_doctypes = [d.name for d in frappe.get_all('DocType')]
                close_doctype = difflib.get_close_matches(user_question, all_doctypes, n=2)
                suggestion = ""
                if close_doctype:
                    suggestion += f"<br>💡 <b>Suggestions:</b> {', '.join(close_doctype)}"
                msg += suggestion
            if 'not permitted' in msg or 'permission' in msg:
                msg += "<br>🔐 <b>Permission issue:</b> You may not have access to this data. Contact your administrator."
            
            result = f"<div class='alert alert-danger'>❌ <b>Error:</b> {msg}</div>"
            add_ai_message(ai_chat, user_question, result, token_usage)
            result_data = {"ai_response": result, "chat_name": ai_chat.name}
            return result_data

    elif intent == "STUDY":
        try:
            # Step 1: Use AI to detect entities and determine if this is a study request
            entity_detection_prompt = [
                {"role": "system", "content": (
                    "Analyze the user's request and determine:\n"
                    "1. Is this a study/analysis request? (true/false)\n"
                    "2. What entities are being analyzed? (item codes, customer names, product names, etc.)\n"
                    "3. What type of analysis is requested? (performance, sales, inventory, etc.)\n\n"
                    "Return a JSON object with:\n"
                    "- 'is_study': boolean\n"
                    "- 'entities': array of entity identifiers found\n"
                    "- 'entity_types': array of entity types (item, customer, product, etc.)\n"
                    "- 'analysis_type': string describing the type of analysis\n"
                    "- 'confidence': float 0-1\n\n"
                    "Examples:\n"
                    "'make a study on item 000125' -> {'is_study': true, 'entities': ['000125'], 'entity_types': ['item'], 'analysis_type': 'item_performance', 'confidence': 0.95}\n"
                    "'analyze customer ABC Corp' -> {'is_study': true, 'entities': ['ABC Corp'], 'entity_types': ['customer'], 'analysis_type': 'customer_analysis', 'confidence': 0.9}\n"
                    "'show me top items' -> {'is_study': false, 'entities': [], 'entity_types': [], 'analysis_type': 'list_query', 'confidence': 0.8}\n"
                    "Be flexible in entity detection - look for any business identifiers, codes, names, or references."
                )},
                {"role": "user", "content": user_question}
            ]
            
            entity_response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=entity_detection_prompt,
                max_tokens=200,
                temperature=0.1,
            )
            
            try:
                entity_analysis = json.loads(entity_response.choices[0].message["content"].strip())
                is_study = entity_analysis.get("is_study", False)
                entities = entity_analysis.get("entities", [])
                entity_types = entity_analysis.get("entity_types", [])
                analysis_type = entity_analysis.get("analysis_type", "")
                confidence = entity_analysis.get("confidence", 0.5)
                
                token_usage["prompt_tokens"] += entity_response['usage']['prompt_tokens']
                token_usage["completion_tokens"] += entity_response['usage']['completion_tokens']
                token_usage["total_tokens"] += entity_response['usage']['total_tokens']
                
            except Exception as e:
                frappe.logger().error(f"Entity detection failed: {str(e)}")
                is_study = False
                entities = []
                entity_types = []
                analysis_type = ""
                confidence = 0.0
            
            # If AI determines this is not a study request, redirect to appropriate handler
            if not is_study or confidence < 0.6:
                frappe.logger().info(f"AI determined not a study request: {user_question} (confidence: {confidence})")
                # Redirect to general data query handler
                try:
                    result = handle_erpnext_module_query("GENERAL_REPORT", user_question, [], confidence, token_usage)
                    add_ai_message(ai_chat, user_question, result, token_usage)
                    result_data = {"ai_response": result, "chat_name": ai_chat.name}
                    return result_data
                except Exception as e:
                    result = ask_enhanced_knowledge_question(chat_history, user_question, "KNOWLEDGE", confidence, token_usage)
                    add_ai_message(ai_chat, user_question, result, token_usage)
                    result_data = {"ai_response": result, "chat_name": ai_chat.name}
                    return result_data
            
            # Use detected entities as keywords
            keywords = entities
            
            if not keywords:
                result = f"<div class='alert alert-warning'>🔍 Could not detect any specific entities for analysis. Please specify what you want to analyze (e.g., 'analyze item 000125' or 'study customer ABC Corp').</div>"
                add_ai_message(ai_chat, user_question, result, token_usage)
                return {"ai_response": result, "chat_name": ai_chat.name}
            
            frappe.logger().info(f"STUDY: Extracted keywords: {keywords}")
            
            # Step 2: Get comprehensive study data
            summary_data = get_item_summary_for_study(keywords)
            
            # Step 3: Use OpenAI to generate a comprehensive analysis
            study_prompt = [
                {"role": "system", "content": (
                    "You are an expert ERP business analyst. Given the following study data (JSON), create a comprehensive business analysis report. "
                    "Include: performance metrics, trends, insights, and actionable recommendations. "
                    "Format as rich HTML with proper styling. Use tables for data, highlight key metrics, and provide clear business insights. "
                    "Focus on practical business value and decision-making support. "
                    "Do NOT use large font sizes. Keep it professional and data-driven."
                )},
                {"role": "user", "content": f"User request: {user_question}\n\nStudy data (JSON):\n{json.dumps(summary_data, indent=2)}"}
            ]
            study_response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=study_prompt,
                max_tokens=1000,
                temperature=0.3,
            )
            result = study_response.choices[0].message["content"].strip()

            # If result is longer than 2000 characters, convert to PDF and return a download link
            if len(result) > 2000:
                if pdfkit is not None:
                    # Create a more professional HTML structure for the PDF
                    html_content = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="UTF-8">
                        <title>Business Analysis Report - {keywords[0] if keywords else 'Study'}</title>
                        <style>
                            body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                            h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
                            h2 {{ color: #34495e; margin-top: 25px; }}
                            table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                            th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                            th {{ background-color: #f2f2f2; font-weight: bold; }}
                            .alert {{ padding: 10px; margin: 10px 0; border-radius: 5px; }}
                            .alert-info {{ background-color: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }}
                            .alert-warning {{ background-color: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }}
                            .alert-danger {{ background-color: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; }}
                        </style>
                    </head>
                    <body>
                        {result}
                    </body>
                    </html>
                    """
                    pdf_file_path = f"/tmp/ai_study_{frappe.generate_hash()}.pdf"
                    pdfkit.from_string(html_content, pdf_file_path)
                    
                    # Read the PDF file content
                    with open(pdf_file_path, "rb") as f:
                        pdf_content = f.read()
                    
                    # Create file document with content
                    file_doc = frappe.get_doc({
                        "doctype": "File",
                        "file_name": f"business_analysis_{keywords[0] if keywords else 'report'}.pdf",
                        "is_private": 0,
                        "content": pdf_content
                    })
                    file_doc.insert(ignore_permissions=True)
                    file_url = file_doc.file_url
                    os.remove(pdf_file_path)
                    result = f"<div class='alert alert-info'>📊 <b>Comprehensive Study Generated!</b><br>📁 <a href='{file_url}' target='_blank' download>Download the full analysis report (PDF)</a><br>📄 <b>Report Length:</b> {len(result):,} characters</div>"
                else:
                    # Fallback to showing first part with note
                    result = result[:2000] + "<br><br><div class='alert alert-info'>📊 <b>Study Summary:</b> This is a condensed version. For the full analysis, please refine your query.</div>"
            
            add_ai_message(ai_chat, user_question, result, token_usage)
            return {"ai_response": result, "chat_name": ai_chat.name}
        except Exception as e:
            msg = str(e)
            frappe.logger().error(f"STUDY error: {msg}")
            result = f"<div class='alert alert-danger'>❌ <b>Study Analysis Error:</b> {msg}<br>💡 Please try specifying the exact item code or customer name.</div>"
            add_ai_message(ai_chat, user_question, result, token_usage)
            return {"ai_response": result, "chat_name": ai_chat.name}



    else:
        # Handle knowledge questions and fallback cases
        try:
            # Use AI to dynamically detect if this might be a study request
            dynamic_study_detection_prompt = [
                {"role": "system", "content": (
                    "Analyze if this request is asking for detailed analysis/study of specific entities. "
                    "Look for:\n"
                    "- Analysis requests (analyze, study, investigate, examine)\n"
                    "- Entity-specific queries (about this item, customer details, product analysis)\n"
                    "- Information gathering (tell me about, what can you tell me, give me details)\n"
                    "- Business intelligence requests (insights, performance, trends)\n\n"
                    "Return JSON: {'is_study': boolean, 'confidence': float 0-1, 'reason': string}\n"
                    "Examples:\n"
                    "'tell me about item 000125' -> {'is_study': true, 'confidence': 0.9, 'reason': 'entity_specific_analysis'}\n"
                    "'show me top customers' -> {'is_study': false, 'confidence': 0.8, 'reason': 'list_query'}\n"
                    "'what is an invoice' -> {'is_study': false, 'confidence': 0.9, 'reason': 'knowledge_question'}"
                )},
                {"role": "user", "content": user_question}
            ]
            
            try:
                study_detection_response = openai.ChatCompletion.create(
                    model="gpt-4",
                    messages=dynamic_study_detection_prompt,
                    max_tokens=100,
                    temperature=0.1,
                )
                
                study_analysis = json.loads(study_detection_response.choices[0].message["content"].strip())
                is_study_request = study_analysis.get("is_study", False)
                study_confidence = study_analysis.get("confidence", 0.5)
                study_reason = study_analysis.get("reason", "")
                
                token_usage["prompt_tokens"] += study_detection_response['usage']['prompt_tokens']
                token_usage["completion_tokens"] += study_detection_response['usage']['completion_tokens']
                token_usage["total_tokens"] += study_detection_response['usage']['total_tokens']
                
            except Exception as e:
                frappe.logger().error(f"Dynamic study detection failed: {str(e)}")
                is_study_request = False
                study_confidence = 0.0
                study_reason = "detection_failed"
            
            if is_study_request and study_confidence > 0.7:
                frappe.logger().info(f"Dynamic study detection: {user_question} (confidence: {study_confidence}, reason: {study_reason})")
                # Use the same AI-driven entity detection as the main STUDY handler
                entity_detection_prompt = [
                    {"role": "system", "content": (
                        "Analyze the user's request and determine:\n"
                        "1. Is this a study/analysis request? (true/false)\n"
                        "2. What entities are being analyzed? (item codes, customer names, product names, etc.)\n"
                        "3. What type of analysis is requested? (performance, sales, inventory, etc.)\n\n"
                        "Return a JSON object with:\n"
                        "- 'is_study': boolean\n"
                        "- 'entities': array of entity identifiers found\n"
                        "- 'entity_types': array of entity types (item, customer, product, etc.)\n"
                        "- 'analysis_type': string describing the type of analysis\n"
                        "- 'confidence': float 0-1\n\n"
                        "Examples:\n"
                        "'make a study on item 000125' -> {'is_study': true, 'entities': ['000125'], 'entity_types': ['item'], 'analysis_type': 'item_performance', 'confidence': 0.95}\n"
                        "'analyze customer ABC Corp' -> {'is_study': true, 'entities': ['ABC Corp'], 'entity_types': ['customer'], 'analysis_type': 'customer_analysis', 'confidence': 0.9}\n"
                        "'show me top items' -> {'is_study': false, 'entities': [], 'entity_types': [], 'analysis_type': 'list_query', 'confidence': 0.8}\n"
                        "Be flexible in entity detection - look for any business identifiers, codes, names, or references."
                    )},
                    {"role": "user", "content": user_question}
                ]
                
                entity_response = openai.ChatCompletion.create(
                    model="gpt-4",
                    messages=entity_detection_prompt,
                    max_tokens=200,
                    temperature=0.1,
                )
                
                try:
                    entity_analysis = json.loads(entity_response.choices[0].message["content"].strip())
                    entities = entity_analysis.get("entities", [])
                    entity_types = entity_analysis.get("entity_types", [])
                    analysis_type = entity_analysis.get("analysis_type", "")
                    
                    token_usage["prompt_tokens"] += entity_response['usage']['prompt_tokens']
                    token_usage["completion_tokens"] += entity_response['usage']['completion_tokens']
                    token_usage["total_tokens"] += entity_response['usage']['total_tokens']
                    
                except Exception as e:
                    frappe.logger().error(f"Entity detection failed in fallback: {str(e)}")
                    entities = []
                    entity_types = []
                    analysis_type = ""
                
                # Use detected entities as keywords
                keywords = entities
                
                if not keywords:
                    result = f"<div class='alert alert-warning'>🔍 Could not detect any specific entities for analysis. Please specify what you want to analyze (e.g., 'analyze item 000125' or 'study customer ABC Corp').</div>"
                    add_ai_message(ai_chat, user_question, result, token_usage)
                    return {"ai_response": result, "chat_name": ai_chat.name}
                
                try:
                    frappe.logger().info(f"STUDY: Extracted keywords: {keywords}")
                    
                    # Step 2: Get comprehensive study data
                    summary_data = get_item_summary_for_study(keywords)
                    
                    # Step 3: Use OpenAI to generate a comprehensive analysis
                    study_prompt = [
                        {"role": "system", "content": (
                            "You are an expert ERP business analyst. Given the following study data (JSON), create a comprehensive business analysis report. "
                            "Include: performance metrics, trends, insights, and actionable recommendations. "
                            "Format as rich HTML with proper styling. Use tables for data, highlight key metrics, and provide clear business insights. "
                            "Focus on practical business value and decision-making support. "
                            "Do NOT use large font sizes. Keep it professional and data-driven."
                        )},
                        {"role": "user", "content": f"User request: {user_question}\n\nStudy data (JSON):\n{json.dumps(summary_data, indent=2)}"}
                    ]
                    study_response = openai.ChatCompletion.create(
                        model="gpt-4",
                        messages=study_prompt,
                        max_tokens=1000,
                        temperature=0.3,
                    )
                    result = study_response.choices[0].message["content"].strip()

                    # If result is longer than 2000 characters, convert to PDF and return a download link
                    if len(result) > 2000:
                        if pdfkit is not None:
                            # Create a more professional HTML structure for the PDF
                            html_content = f"""
                            <!DOCTYPE html>
                            <html>
                            <head>
                                <meta charset="UTF-8">
                                <title>Business Analysis Report - {keywords[0] if keywords else 'Study'}</title>
                                <style>
                                    body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
                                    h1 {{ color: #2c3e50; border-bottom: 2px solid #3498db; padding-bottom: 10px; }}
                                    h2 {{ color: #34495e; margin-top: 25px; }}
                                    table {{ border-collapse: collapse; width: 100%; margin: 15px 0; }}
                                    th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                                    th {{ background-color: #f2f2f2; font-weight: bold; }}
                                    .alert {{ padding: 10px; margin: 10px 0; border-radius: 5px; }}
                                    .alert-info {{ background-color: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; }}
                                    .alert-warning {{ background-color: #fff3cd; border: 1px solid #ffeaa7; color: #856404; }}
                                    .alert-danger {{ background-color: #f8d7da; border: 1px solid #f5c6cb; color: #721c24; }}
                                </style>
                            </head>
                            <body>
                                {result}
                            </body>
                            </html>
                            """
                            pdf_file_path = f"/tmp/ai_study_{frappe.generate_hash()}.pdf"
                            pdfkit.from_string(html_content, pdf_file_path)
                            
                            # Read the PDF file content
                            with open(pdf_file_path, "rb") as f:
                                pdf_content = f.read()
                            
                            # Create file document with content
                            file_doc = frappe.get_doc({
                                "doctype": "File",
                                "file_name": f"business_analysis_{keywords[0] if keywords else 'report'}.pdf",
                                "is_private": 0,
                                "content": pdf_content
                            })
                            file_doc.insert(ignore_permissions=True)
                            file_url = file_doc.file_url
                            os.remove(pdf_file_path)
                            result = f"<div class='alert alert-info'>📊 <b>Comprehensive Study Generated!</b><br>📁 <a href='{file_url}' target='_blank' download>Download the full analysis report (PDF)</a><br>📄 <b>Report Length:</b> {len(result):,} characters</div>"
                        else:
                            # Fallback to showing first part with note
                            result = result[:2000] + "<br><br><div class='alert alert-info'>📊 <b>Study Summary:</b> This is a condensed version. For the full analysis, please refine your query.</div>"
                    
                    add_ai_message(ai_chat, user_question, result, token_usage)
                    return {"ai_response": result, "chat_name": ai_chat.name}
                except Exception as e:
                    msg = str(e)
                    frappe.logger().error(f"STUDY error: {msg}")
                    result = f"<div class='alert alert-danger'>❌ <b>Study Analysis Error:</b> {msg}<br>💡 Please try specifying the exact item code or customer name.</div>"
                    add_ai_message(ai_chat, user_question, result, token_usage)
                    return {"ai_response": result, "chat_name": ai_chat.name}
            
            # Check if this might be a data query despite intent classification
            question_lower = user_question.lower()
            data_keywords = ['top', 'list', 'show', 'get', 'find', 'count', 'total', 'report', 'data', 'items', 'customers', 'sales', 'purchase', 'this year', 'this month']
            
            # More aggressive fallback for data queries
            if any(keyword in question_lower for keyword in data_keywords):
                frappe.logger().info(f"Attempting SQL generation as fallback for: {user_question}")
                # Try multiple approaches
                sql_query = None
                
                # Try with STOCK intent first for item-related queries
                if any(word in question_lower for word in ['item', 'items', 'product', 'products']):
                    sql_query = generate_enhanced_sql(user_question, "STOCK", ["Item", "Sales Invoice Item"], token_usage)
                
                # Try with SELLING intent for sales-related queries
                elif any(word in question_lower for word in ['sales', 'customer', 'customers', 'invoice']):
                    sql_query = generate_enhanced_sql(user_question, "SELLING", ["Sales Invoice", "Customer"], token_usage)
                
                # Try with GENERAL_REPORT as last resort
                if not sql_query:
                    sql_query = generate_enhanced_sql(user_question, "GENERAL_REPORT", [], token_usage)
                
                # If all AI-generated SQL fails, try direct SQL for common patterns
                if not sql_query and 'top' in question_lower and 'items' in question_lower:
                    frappe.logger().info("Using direct SQL generation for top items query")
                    current_year = datetime.now().year
                    sql_query = f"""
                    SELECT 
                        i.item_code,
                        i.item_name,
                        SUM(sii.qty) as total_qty,
                        SUM(sii.amount) as total_amount
                    FROM tabItem i
                    JOIN `tabSales Invoice Item` sii ON i.item_code = sii.item_code
                    JOIN `tabSales Invoice` si ON sii.parent = si.name
                    WHERE si.docstatus = 1 
                    AND YEAR(si.posting_date) = {current_year}
                    GROUP BY i.item_code, i.item_name
                    ORDER BY total_amount DESC, total_qty DESC
                    LIMIT 10
                    """
                
                if sql_query:
                    frappe.logger().info(f"Generated SQL: {sql_query}")
                    db_result = frappe.db.sql(sql_query, as_dict=True)
                    if db_result:
                        if len(db_result) > 10 or len(db_result[0].keys()) > 5:
                            result = generate_excel_file(db_result)
                        else:
                            result = polish_erp_answer_html(user_question, format_result(db_result), token_usage)
                    else:
                        result = f"<div class='alert alert-info'>🔍 No data found for your query. Try adjusting your criteria.</div>"
                else:
                    result = ask_enhanced_knowledge_question(chat_history, user_question, intent, confidence, token_usage)
            else:
                result = ask_enhanced_knowledge_question(chat_history, user_question, intent, confidence, token_usage)
            
            add_ai_message(ai_chat, user_question, result, token_usage)
            result_data = {"ai_response": result, "chat_name": ai_chat.name}
            
            # Cache knowledge questions longer (static content)
            cache_expiry = determine_cache_expiry(user_question, intent, suggested_doctypes)
            if cache_expiry > 0:
                set_cached_response(cache_key, result_data, cache_expiry)
            
            return result_data
        except Exception as e:
            msg = str(e)
            result = f"<div class='alert alert-danger'>❌ <b>Unexpected error:</b> {msg}<br>💬 Please try rephrasing your question or contact support.</div>"
            add_ai_message(ai_chat, user_question, result, token_usage)
            result_data = {"ai_response": result, "chat_name": ai_chat.name}
            return result_data


def preprocess_question(current_question: str, chat_history: list) -> str:
    current_question = current_question.strip()
    
    # Enhanced context detection for follow-up questions
    follow_up_indicators = [
        r"^(by|include|only|group by|filter|sort by|add|show|with|also|and|plus|including|exclude|without)",
        r"^(top|bottom|first|last|latest|recent|oldest)",
        r"^(this|that|these|those|it|them)",
        r"^(more|less|higher|lower|bigger|smaller)",
        r"^(today|yesterday|tomorrow|this week|this month|this year)",
        r"^(customer|supplier|item|product|invoice|order|payment)",
        r"^(amount|total|sum|count|average|max|min)",
        r"^(outstanding|overdue|pending|completed|cancelled|draft)"
    ]
    
    # Check if this looks like a follow-up question
    is_follow_up = any(re.match(pattern, current_question, re.I) for pattern in follow_up_indicators)
    
    if is_follow_up or len(current_question.split()) < 6:
        # Look for the most recent user question that's substantial
        for msg in reversed(chat_history):
            if msg.get("role") == "user" and len(msg["content"].split()) >= 4:
                previous_question = msg["content"].strip()
                # Avoid infinite loops by checking if we're not already combining
                if not any(indicator in previous_question.lower() for indicator in ["include", "add", "with", "also", "and", "plus"]):
                    combined_question = f"{previous_question} {current_question}"
                    frappe.logger().info(f"Combined question: '{previous_question}' + '{current_question}' = '{combined_question}'")
                    return combined_question
                break
    
    return current_question



def generate_sql_from_question(question: str, token_usage=None) -> Optional[str]:
    if token_usage is None:
        token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    messages = [
    {
        "role": "system",
        "content": (
            "You are an expert assistant that converts complex natural language questions into secure, valid, and syntactically correct "
            "SQL SELECT queries for ISOFT ERP, a system built on ERPNext.\n\n"

            " NEVER return explanations. Only return a valid SQL SELECT query.\n"
            " Only generate SELECT queries — NO INSERT, UPDATE, DELETE, or DDL.\n\n"

            " ERP STRUCTURE AND CONVENTIONS:\n"
            "- Tables follow the ERPNext format: all table names are prefixed with `tab`.\n"
            "- For transactional documents like Sales Invoice, Purchase Invoice, Delivery Note: filter with `docstatus = 1` (submitted).\n"
            "- For Stock Ledger Entry, GL Entry, and other tables with `is_cancelled` field, filter with `is_cancelled = 0`.\n"
            "- Some tables like `tabBin` do not have `docstatus` or `is_cancelled`.\n"
            "- Parent and child documents are linked using `parent`, `parenttype`, and `parentfield`.\n"
            "- Use `JOIN` for parent-child queries (e.g., Invoice + Invoice Items).\n"
            "- Use subqueries when you need filtering on aggregates, last values, or conditional logic.\n\n"

            " STOCK MANAGEMENT:\n"
            "- `tabBin`: current stock — `item_code`, `warehouse`, `actual_qty`, `valuation_rate`, `reserved_qty`, etc.\n"
            "- `tabStock Ledger Entry`: detailed movements — `posting_date`, `item_code`, `actual_qty`, `stock_value_difference`, `voucher_type`, `voucher_no`, etc.\n"
            "- For valuation, use: `actual_qty * valuation_rate`.\n\n"
            "- `tabItem Price`: item price — `item_code`, `uom`, `selling`, `buying`, `price_list`, `price_list_rate`, `creation`, etc. it's not has posting_date"

            " SALES MODULE:\n"
            "- `tabSales Invoice`: `customer`, `grand_total`, `posting_date`, `is_return`, `return_against`.\n"
            "- `tabSales Invoice Item`: `item_code`, `qty`, `rate`, `amount`, `parent`.\n"
            "- `is_return = 1` means this is a return or credit note (negative invoice).\n\n"

            " PURCHASE MODULE:\n"
            "- `tabPurchase Invoice` and `tabPurchase Receipt`: `supplier`, `posting_date`, `grand_total`, `is_return`, `return_against`.\n"
            "- `is_return = 1` indicates a returned or reversed purchase.\n"
            "- Use `return_against` to trace the original document.\n\n"

            " DELIVERY MODULE:\n"
            "- `tabDelivery Note`: `customer`, `posting_date`, `is_return`, `return_against`.\n"
            "- If `is_return = 1`, this is a returned delivery.\n\n"

            " ACCOUNTING:\n"
            "- `tabGL Entry`: `posting_date`, `account`, `debit`, `credit`, `voucher_type`, `party`.\n"
            "- Use to trace financial transactions from all modules.\n\n"

            " HR & PAYROLL:\n"
            "- `tabEmployee`, `tabSalary Slip`: `net_pay`, `employee`, `posting_date`, `salary_structure`.\n\n"

            " MANUFACTURING:\n"
            "- `tabWork Order`: `production_item`, `qty`, `status`, `actual_start_date`, `actual_end_date`.\n"
            "- `tabBOM`, `tabBOM Item`: `item_code`, `qty`, `rate`.\n\n"

            " QUERY LOGIC GUIDELINES:\n"
            "- Use `JOIN` when pulling from both parent and child tables.\n"
            "- Use `GROUP BY`, `SUM()`, `HAVING`, etc. when aggregating.\n"
            "- Use `subqueries` to find latest entries, filter by conditions, or compute derived values.\n"
            "- Use table aliasing for clarity (e.g., `si`, `sii`, `pi`, `pri`, `dn`, `dnl`, etc.).\n"
            "- Always apply `is_return` filter when analyzing net sales/purchases/deliveries.\n\n"

            " EXAMPLES:\n"
            "Q: Total net sales (excluding returns) per customer this year\n"
            "A: SELECT customer, SUM(grand_total) AS net_sales FROM `tabSales Invoice` WHERE docstatus = 1 AND is_return = 0 AND posting_date >= '2025-01-01' GROUP BY customer\n\n"

            "Q: Items returned to suppliers in June 2025\n"
            "A: SELECT pri.item_code, SUM(pri.qty) AS total_returned FROM `tabPurchase Receipt Item` pri JOIN `tabPurchase Receipt` pr ON pri.parent = pr.name WHERE pr.docstatus = 1 AND pr.is_return = 1 AND pr.posting_date BETWEEN '2025-06-01' AND '2025-06-30' GROUP BY pri.item_code\n\n"

            "Q: Get the original invoice number for returned sales invoices\n"
            "A: SELECT name, return_against FROM `tabSales Invoice` WHERE docstatus = 1 AND is_return = 1\n\n"

            "Now convert the following question into a valid SQL SELECT query:"
        )
    },
    {
        "role": "user",
        "content": question
    }
]


    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        max_tokens=300,
        temperature=0,
    )

    token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += response['usage']['completion_tokens']
    token_usage["total_tokens"] += response['usage']['total_tokens']
    
    sql = (response.choices[0].message['content'].strip()).rstrip(";")
    if not sql.lower().startswith("select") or re.search(r";|insert|update|delete|drop|alter|truncate", sql, re.I):
        return None

    # Validate SQL fields using metadata
    validation_error = validate_sql_fields(sql)
    if validation_error:
        frappe.throw(f"SQL validation error: {validation_error}")

    return sql


def validate_sql_fields(sql: str) -> Optional[str]:
    """
    Parses the SQL SELECT query, extracts table and field names, and checks them against DocType metadata.
    Returns None if valid, or an error message string if invalid.
    """
    if sqlparse is None:
        return "SQL validation requires the 'sqlparse' library. Please install it."
    try:
        parsed = sqlparse.parse(sql)
        if not parsed:
            return "Could not parse SQL."
        stmt = parsed[0]
        # Find all table names (FROM, JOIN)
        tables = set()
        fields = set()
        from_seen = False
        for token in stmt.tokens:
            if token.ttype is None and token.is_group:
                for sub in token.tokens:
                    if sub.ttype is None and sub.is_group:
                        for subsub in sub.tokens:
                            if subsub.ttype is None and subsub.is_group:
                                continue
                            if subsub.value.upper() == 'FROM' or subsub.value.upper() == 'JOIN':
                                from_seen = True
                            elif from_seen and subsub.ttype is None:
                                table_name = subsub.value.strip('`')
                                if table_name.startswith('tab'):
                                    tables.add(table_name)
                                from_seen = False
            # Find fields in SELECT
            if token.ttype is None and token.is_group:
                for sub in token.tokens:
                    if sub.ttype is None and sub.is_group:
                        for subsub in sub.tokens:
                            if subsub.ttype is None and subsub.is_group:
                                continue
                            if subsub.value.upper() == 'SELECT':
                                continue
                            if subsub.value.upper() == 'FROM':
                                break
                            if subsub.ttype is None:
                                field_expr = subsub.value.strip()
                                if '.' in field_expr:
                                    fields.add(field_expr.split('.')[-1].strip('`, '))
                                else:
                                    fields.add(field_expr.strip('`, '))
        # Validate tables and fields
        for table in tables:
            doctype = table[3:] if table.startswith('tab') else table
            try:
                meta = frappe.get_meta(doctype)
            except Exception:
                return f"Table {table} (DocType {doctype}) does not exist."
            valid_fields = {f.fieldname for f in meta.fields}
            valid_fields.add('name')
            for field in fields:
                if field not in valid_fields:
                    return f"Field '{field}' does not exist in {doctype}."
        return None
    except Exception as e:
        return f"SQL validation error: {str(e)}"


def format_result(results: list) -> str:
    max_rows = 10
    limited_results = results[:max_rows]

    formatted_rows = []
    for row in limited_results:
        formatted_row = ", ".join(f"{escape_html(str(k))}: {escape_html(str(v))}" for k, v in row.items())
        formatted_rows.append(formatted_row)

    if len(results) > max_rows:
        formatted_rows.append(f"... (truncated {len(results) - max_rows} more rows)")

    return "\n".join(formatted_rows)


def generate_excel_file(results: list) -> str:
    if not results:
        frappe.throw("No results to export.")

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return generate_csv_file(results)

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "ERP Query Results"

        # Get column headers from the first result
        headers = list(results[0].keys())
        
        # Write headers with formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            if cell is not None:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

        # Write data rows
        for row_num, row_data in enumerate(results, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                
                # Handle different data types
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    # Keep numbers as numbers
                    pass
                elif isinstance(value, str) and value.isdigit() and value.startswith("0"):
                    # Preserve leading zeros
                    value = f'="{value}"'
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if cell is not None:
                    # Add borders and alignment
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            if column and column[0] and column[0].column is not None:
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell and cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                # Set column width (min 10, max 50)
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        # Save to temporary file
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()

        # Read the file and create Frappe File document
        with open(temp_file.name, 'rb') as f:
            file_content = f.read()

        # Clean up temporary file
        import os
        os.unlink(temp_file.name)

        # Create file document
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": "erp_query_result.xlsx",
            "content": file_content,
            "is_private": 0
        }).insert(ignore_permissions=True)

        return file_doc.file_url
    else:
        # Fallback to CSV if worksheet creation failed
        return generate_csv_file(results)

def generate_csv_file(results: list) -> str:
    """Fallback function for CSV generation when Excel is not available"""
    if not results:
        frappe.throw("No results to export.")

    normalized_results = []
    for row in results:
        plain_row = {}
        for k, v in row.items():
            if v is None:
                plain_row[k] = ""
            elif isinstance(v, str) and v.isdigit() and v.startswith("0"):
                # Preserve leading zeros in Excel
                plain_row[k] = f'="{v}"'
            else:
                plain_row[k] = str(v)
        normalized_results.append(plain_row)

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=normalized_results[0].keys())
    writer.writeheader()
    writer.writerows(normalized_results)
    csv_content = output.getvalue()
    output.close()

    file_doc = frappe.get_doc({
        "doctype": "File",
        "file_name": "erp_query_result.csv",
        "content": csv_content,
        "is_private": 0
    }).insert(ignore_permissions=True)

    return file_doc.file_url



def polish_erp_answer_html(question: str, db_result: str, token_usage=None) -> str:
    if token_usage is None:
        token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    max_len = 600  # Lowered from 1000 for efficiency
    truncated_result = db_result[:max_len] + "\n... (truncated)" if len(db_result) > max_len else db_result
    # --- SHORTER SYSTEM PROMPT FOR ERP FORMATTING ---
    messages = [
        {
            "role": "system",
            "content": (
                "Format ERP query results as readable text table."
            )
        },
        {
            "role": "user",
            "content": f"Q: {question}\nResult:\n{truncated_result}"
        }
    ]
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        max_tokens=700,  # Lowered from 1000 for efficiency
        temperature=0.3,
    )
    token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += response['usage']['completion_tokens']
    token_usage["total_tokens"] += response['usage']['total_tokens']
    return response.choices[0].message['content'].strip()


def ask_knowledge_question_html(chat_history, current_question: str, token_usage=None) -> str:
    if token_usage is None:
        token_usage = {"prompt_tokens": 0, "completion_tokens": 0, "total_tokens": 0}
    # --- LIMIT CHAT HISTORY TO LAST 2 USER/ASSISTANT TURNS FOR TOKEN EFFICIENCY ---
    trimmed_history = [msg for msg in chat_history if msg.get("role") == "system"]
    # Get last 2 user/assistant turns (4 messages max)
    turns = [msg for msg in chat_history if msg.get("role") in ("user", "assistant")]
    trimmed_history.extend(turns[-4:])
    if not any(msg.get("role") == "system" for msg in trimmed_history):
        trimmed_history.insert(0, {
            "role": "system",
            "content": (
                "You are a helpful assistant for ISOFT ERP, knowledgeable in business, accounting, and ERP usage. "
                "Do not mention 'erpnext v13'."
            )
        })
    trimmed_history.append({"role": "user", "content": current_question})
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=trimmed_history,
        max_tokens=700,  # Lowered from 1000 for efficiency
        temperature=0.3,
    )
    token_usage["prompt_tokens"] += response['usage']['prompt_tokens']
    token_usage["completion_tokens"] += response['usage']['completion_tokens']
    token_usage["total_tokens"] += response['usage']['total_tokens']
    return response.choices[0].message['content'].strip()

def clean_intent(raw_intent: str) -> str:
    cleaned = raw_intent.strip().strip("'\"").upper()
    valid_intents = ("ERP", "KNOWLEDGE", "CLARIFY", "STUDY", "SUPPLIER_STUDY", "PROJECT_ANALYSIS", "PAYROLL_SUMMARY", "ASSET_ANALYSIS", "HR_SUMMARY")
    if cleaned not in valid_intents:
        cleaned = "KNOWLEDGE"
    return cleaned


def generate_ai_chat_title(first_message: str) -> str:
    """Generate a concise AI chat title based on the first user message using OpenAI."""
    response = openai.ChatCompletion.create(
        model="gpt-4",
        messages=[
            {"role": "system", "content": "Generate a short, clear chat title for this user message. Do not use quotes."},
            {"role": "user", "content": first_message}
        ],
        max_tokens=12,
        temperature=0.2,
    )
    return response.choices[0].message["content"].strip()

def get_or_create_ai_chat(ai_chat_name: str = "", first_message: str = "") -> Document:
    """
    If ai_chat_name is provided and exists, return that chat.
    Otherwise, create a new AI Chat with a generated title based on the first message.
    """
    if ai_chat_name:
        try:
            return frappe.get_doc("AI Chat", ai_chat_name)
        except Exception:
            pass  # If not found, fall through to create new
    # Generate title if not provided
    title = generate_ai_chat_title(first_message) if first_message else "AI Chat"
    doc = frappe.new_doc("AI Chat")
    doc.title = title
    doc.owner = frappe.session.user
    doc.insert(ignore_permissions=True)
    return doc


def add_ai_message(ai_chat, user_question, ai_response, token_usage):
    ai_chat.append("messages", {
        "user_question": user_question,
        "ai_response": ai_response,
        "prompt_tokens": token_usage["prompt_tokens"],
        "completion_tokens": token_usage["completion_tokens"],
        "total_tokens": token_usage["total_tokens"]
    })
    ai_chat.save()

@frappe.whitelist()
def get_user_ai_chats(owner_id: str) -> List[Dict]:
    if frappe.session.user == "Guest":
        return []
    """
    Return all AI Chat documents for a given owner id, excluding soft deleted.
    Returns a list of dicts with 'name' and 'title'.
    """
    return frappe.get_all(
        "AI Chat",
        filters={"owner": owner_id, "soft_delete": ["!=", 1]},
        fields=["name", "title"],
        order_by="creation desc"
    )


@frappe.whitelist()
def get_ai_chat_messages(chat_name: str) -> List[Dict]:
    if frappe.session.user == "Guest":
        return []
    """
    Return all messages for a given AI Chat document as a list of dicts.
    """
    doc = frappe.get_doc('AI Chat', chat_name)
    return [msg.as_dict() for msg in doc.messages]

@frappe.whitelist()
def soft_delete_ai_chat(chat_name: str):
    if frappe.session.user == "Guest":
        return {"success": False}
    """
    Soft delete an AI Chat by setting soft_delete=1.
    """
    try:
        doc = frappe.get_doc('AI Chat', chat_name)
        doc.soft_delete = 1
        doc.save(ignore_permissions=True)
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}


@frappe.whitelist()
def get_item_summary_for_study(keywords: List[str]) -> Dict:
    if frappe.session.user == "Guest":
        return {}
    if not keywords:
        frappe.throw("Keywords array cannot be empty.")

    # Dynamically get Item fields (including custom fields)
    item_meta = frappe.get_meta('Item')
    item_fields = [f.fieldname for f in item_meta.fields if f.fieldtype not in ('Section Break', 'Column Break', 'Tab Break')] + ['name']
    # Use a default set if some fields are missing
    default_fields = ['item_code', 'item_name', 'brand', 'item_group', 'description']
    select_fields = [f for f in default_fields if f in item_fields]
    if not select_fields:
        select_fields = [item_fields[0]]  # fallback to at least one field

    # Prepare the WHERE clause dynamically
    like_clauses = []
    for kw in keywords:
        kw_escaped = frappe.db.escape('%' + kw + '%')
        like_parts = []
        for field in select_fields:
            like_parts.append(f"i.{field} LIKE {kw_escaped}")
        like_clauses.append('(' + ' OR '.join(like_parts) + ')')
    where_clause = " OR ".join(like_clauses)

    # Build SELECT clause
    select_sql = ',\n        '.join([f'i.{f}' for f in select_fields])
    # Add analytics fields (static for now, can be made dynamic per module)
    select_sql += ",\n        DATE_FORMAT(COALESCE(q.transaction_date, si.posting_date, pi.posting_date), '%Y-%m') AS month_year,\n\n        COUNT(DISTINCT q.name) AS quotation_count,\n        IFNULL(SUM(qi.qty), 0) AS quotation_items_qty,\n\n        COUNT(DISTINCT si.name) AS sales_invoice_count,\n        IFNULL(SUM(sii.qty), 0) AS sales_invoice_items_qty,\n\n        COUNT(DISTINCT pi.name) AS purchase_invoice_count,\n        IFNULL(SUM(pii.qty), 0) AS purchase_invoice_items_qty"

    query = f"""
    SELECT
        {select_sql}
    FROM tabItem i

    LEFT JOIN `tabQuotation Item` qi ON qi.item_code = i.item_code
    LEFT JOIN `tabQuotation` q ON q.name = qi.parent AND q.docstatus = 1

    LEFT JOIN `tabSales Invoice Item` sii ON sii.item_code = i.item_code
    LEFT JOIN `tabSales Invoice` si ON si.name = sii.parent AND si.docstatus = 1

    LEFT JOIN `tabPurchase Invoice Item` pii ON pii.item_code = i.item_code
    LEFT JOIN `tabPurchase Invoice` pi ON pi.name = pii.parent AND pi.docstatus = 1

    WHERE {where_clause}
    GROUP BY i.item_code, month_year
    ORDER BY month_year DESC
    """

    items = frappe.db.sql(query, as_dict=True)
    item_codes = list({item.get('item_code', item.get('name')) for item in items})
    stock_by_item = {}
    if item_codes:
        stock_rows = frappe.db.sql(f"""
            SELECT item_code, warehouse, actual_qty
            FROM tabBin
            WHERE item_code IN ({', '.join(['%s']*len(item_codes))})
        """, item_codes, as_dict=True)
        for row in stock_rows:
            stock_by_item.setdefault(row['item_code'], []).append({
                'warehouse': row['warehouse'],
                'actual_qty': row['actual_qty']
            })
    for item in items:
        item['stock_by_warehouse'] = stock_by_item.get(item.get('item_code', item.get('name')), [])
    return {'items': items}


def get_child_tables_for_parent(parent_doctype: str):
    """
    Returns a list of (child_table, link_field) for all child tables of the given parent DocType.
    """
    meta = frappe.get_meta(parent_doctype)
    child_tables = []
    for field in meta.fields:
        if field.fieldtype == 'Table' and field.options:
            # The link field in the child table is always 'parent' (ERPNext convention)
            child_tables.append((field.options, 'parent'))
    return child_tables


